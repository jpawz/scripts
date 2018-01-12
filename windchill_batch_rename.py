#! python3
# Batch renames WTparts and CADdocuments in Windchill
# Input file (*.xlsx) should contain three columns:
#      part_number | long_name | short_name
# as processing goes by fourth column will be filled with statuses:
# ok - rename successful, fail - rename failed, empty - not yet processed.

import openpyxl
import sys
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

if len(sys.argv) != 2:
    print("use: windchill_batch_rename.py excel_file.xlsx")
    exit()

xlsx_file = sys.argv[1]
workbook = load_workbook(filename=xlsx_file)
sheet = workbook.active
windchill_url = "http://windchill_url"
browser = webdriver.Chrome()
browser.set_page_load_timeout(5)
browser.get(windchill_url)
main_window = browser.current_window_handle
input("login and press [Enter] to continue...")


def find_part(part_number):
    find_field = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located(By.ID, "gloabalSearchField")
    )
    find_field.clear()
    find_field.send_keys(part_number)
    find_field.send_keys(Keys.ENTER)


def open_rename_window(part_number):
    number_link = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located(By.LINK_TEXT, part_number)
    )
    ac = ActionChains(browser)
    ac.context_click(number_link).perform()
    rename_link = WebDriverWait(browser, 2).until(
        EC.presence_of_element_located(By.LINK_TEXT, "Rename")
    )
    rename_link.click()


def handle_rename_window(new_name):
    WebDriverWait(browser, 5).until(EC.new_window_is_opened)
    browser.switch_to.window(browser.window_handles[1])  # new window
    try:
        wait = WebDriverWait(browser, 3)
        rename_field = wait.until(
            EC.presence_of_element_located(
                By.XPATH, "//input[contains(@id, 'NameInputId')]"
            )
        )
        rename_field.clear()
        rename_field.send_keys(new_name)
        ok_btn = browser.find_element_by_xpath(
            "//button[contains(@accesskey, 'o')]")  # c - cancel, o - OK
        ok_btn.click()
    except:
        raise
    finally:
        browser.close()
        browser.switch_to.window(main_window)


def rename_wtpart(part_number, new_name):
    browser.get(windchill_url)
    find_part(part_number)
    open_rename_window(part_number)
    handle_rename_window(new_name)


for row in sheet.iter_rows(max_col=4):
    try:
        rename_wtpart(row[0], row[2])
        row[3] = "OK"  # if success
    except Exception as ex:
        print("error at " + row[0])
        print(ex)
        continue
    workbook.save(xlsx_file)
