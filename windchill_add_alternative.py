#! python3
# Batch adds alternative to WTparts in Windchill
# Input file (*.xlsx) with three columns:
#	  part_number | alternative_part_number | two_way (Yes/No)
# as processing goes by, fourth column will be filled with statuses.

import sys

import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

if len(sys.argv) != 2:
    print("use: skrypt.py excel_file.xlsx")
    exit()


class MissingNumberError(NoSuchElementException):
    '''raise when number was not found'''


xlsx_file = sys.argv[1]
workbook = load_workbook(filename=xlsx_file)
sheet = workbook.active
windchill_url = ""

cap = DesiredCapabilities().INTERNETEXPLORER
cap['ignoreProtectedModeSettings'] = True
cap['IntroduceInstabilityByIgnoringProtectedModeSettings'] = True
cap['nativeEvents'] = True
cap['ignoreZoomSetting'] = True
cap['requireWindowFocus'] = True
cap['INTRODUCE_FLAKINESS_BY_IGNORING_SECURITY_DOMAINS'] = True

browser = webdriver.Ie(
    capabilities=cap,
    executable_path=r'D:\WinPython\python\IEDriverServer.exe')
browser.set_page_load_timeout(5)
browser.get(windchill_url)
input("login and press [Enter] to continue...")
main_window = browser.current_window_handle
wait_5_sec = WebDriverWait(browser, 5)
MSG_MISSING_NR = "NUMBER NOT FOUND"
MSG_MISSING_ALT_NR = "ALT NUMBER NOT FOUND"


def find_part(part_number):
    browser.get(windchill_url)
    find_field = wait_5_sec.until(
        EC.presence_of_element_located((By.ID, "gloabalSearchField")))
    find_field.clear()
    find_field.send_keys(part_number)
    find_field.send_keys(Keys.ENTER)


def open_part_info_window(part_number):
    try:
        number_link = wait_5_sec.until(
            EC.presence_of_element_located((By.PARTIAL_LINK_TEXT,
                                            part_number)))
    except TimeoutException:
        raise MissingNumberError(MSG_MISSING_NR)
    else:
        number_link.click()


def switch_to_tab(link_text):
    tab = wait_5_sec.until(
        EC.presence_of_element_located((By.LINK_TEXT, link_text)))
    tab.click()


def alternative_is_already_added(alternative_number):
    alternates_table = wait_5_sec.until(
        EC.presence_of_element_located((By.ID, "netmarkets.alternates.list")))
    alt_nums = alternates_table.find_elements_by_link_text(alternative_number)
    if len(alt_nums) == 1:
        return True
    else:
        return False


def open_find_alternate_part_window():
    alt_bar = wait_5_sec.until(
        EC.presence_of_element_located((By.ID,
                                        "netmarkets.alternates.list.toolBar")))
    alt_bar.location_once_scrolled_into_view
    buttons = alt_bar.find_elements_by_tag_name("button")
    add_button = buttons[0]
    add_button.click()


def handle_find_alternative_part_window(alternative_number):
    wait_5_sec.until(EC.number_of_windows_to_be(2))
    if (main_window == browser.window_handles[0]):
        new_window = browser.window_handles[1]
    else:
        new_window = browser.window_handles[0]
    browser.switch_to.window(new_window)
    try:
        find_number_field = wait_5_sec.until(
            EC.presence_of_element_located((By.ID, "number2_SearchTextBox")))
        find_number_field.clear()
        find_number_field.send_keys(alternative_number)
        find_number_field.send_keys(Keys.ENTER)
        search_button = browser.find_element_by_name("pickerSearch")
        search_button.click()
        checker = wait_5_sec.until(
            EC.presence_of_element_located((By.CLASS_NAME,
                                            "x-grid3-col-checker")))
        tab_rows = browser.find_elements_by_class_name("x-grid3-row-table")
        if (len(tab_rows) > 1):
            raise ValueError("multiple elements found")
    except TimeoutException:
        browser.switch_to_window(new_window)
        browser.close()
        raise MissingNumberError(MSG_MISSING_ALT_NR)
    else:
        checker.click()
        ok_btn = browser.find_element_by_id("ext-gen40")
        ok_btn.click()
        try:
            WebDriverWait(browser, 1).until(EC.alert_is_present())
            browser.switch_to.alert.accept()
            browser.switch_to_window(new_window)
            browser.close()
        except TimeoutException:
            pass
    finally:
        browser.switch_to.window(main_window)


def handle_edit_alternate_window(two_way):
    wait_5_sec.until(EC.number_of_windows_to_be(2))
    if (main_window == browser.window_handles[0]):
        new_window = browser.window_handles[1]
    else:
        new_window = browser.window_handles[0]
    browser.switch_to.window(new_window)
    wait_5_sec.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
    select = Select(browser.find_element_by_tag_name("select"))
    if two_way == "Yes":
        select.select_by_value("True")
    else:
        select.select_by_value("False")
    ok_btn = browser.find_element_by_id("ext-gen36")
    ok_btn.click()
    browser.switch_to.window(main_window)


def check_settings_for_alternative(alternative_number, two_way):
    table = browser.find_element_by_id(
        "netmarkets.alternates.list").find_element_by_class_name(
            "x-grid3-body")
    rows = table.find_elements_by_tag_name("table")
    numbers = table.find_elements_by_link_text(alternative_number)
    if not rows or not numbers:
        raise ValueError("can't find alternate")
    else:
        for i in range(len(rows)):
            if rows[i].find_elements_by_link_text(alternative_number):
                tw = rows[i].find_element_by_class_name(
                    "x-grid3-col-twoWay").text
                if (tw == two_way):
                    return
                else:
                    checker = rows[i].find_element_by_class_name(
                        "x-grid3-row-checker")
                    checker.click()
                    alt_bar = browser.find_element_by_id(
                        "netmarkets.alternates.list.toolBar")
                    buttons = alt_bar.find_elements_by_tag_name("button")
                    edit_button = buttons[2]
                    edit_button.click()
                    handle_edit_alternate_window(two_way)
                    break


def add_alternative_part(part_number, alt_number, two_way):
    browser.switch_to.window(main_window)
    find_part(part_number)
    open_part_info_window(part_number)
    switch_to_tab("xxx")
    if (alternative_is_already_added(alt_number)):
        check_settings_for_alternative(alt_number, two_way)
    else:
        open_find_alternate_part_window()
        handle_find_alternative_part_window(alt_number)
        check_settings_for_alternative(alt_number, two_way)


def part_is_already_checked():
    if ((str(row[3].value) == MSG_MISSING_NR)
            or (str(row[3].value) == MSG_MISSING_ALT_NR)
            or (str(row[3].value) == "OK")):
        return True
    return False


def close_all_but_main_windows():
    for window in browser.window_handles:
        if window != main_window:
            browser.switch_to_window(window)
            browser.close()
    browser.switch_to_window(main_window)


for index, row in enumerate(sheet.rows):
    print(index)
    part_n = str(row[0].value).strip()
    alt_n = str(row[1].value).strip()
    two_way = str(row[2].value).strip()
    if ((not part_n) or (not alt_n) or (not two_way) or (part_n == "None")
            or (alt_n == "None") or (two_way == "None")):
        continue
    if (part_is_already_checked()):
        continue
    error_msg = ""
    for attempt in range(3):
        try:
            add_alternative_part(part_n, alt_n, two_way)
            sheet.cell(row=index + 1, column=4).value = "OK"
            break
        except MissingNumberError as ex:
            sheet.cell(row=index + 1, column=4).value = str(ex)
            break
        except Exception as ex:
            error_msg = str(ex)
            close_all_but_main_windows()
    else:
        sheet.cell(row=index + 1, column=4).value = error_msg
    print(error_msg)
    workbook.save(xlsx_file)
workbook.save(xlsx_file)
browser.quit()
