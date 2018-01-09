#! python3
# Scraper: get short names from Products Catalogue website
# Input: xlsx file with first column containing part numbers (continous)
#        and optionally long names in second column
# Output: the same xlsx file with third column filled with
#         short names

import sys
import openpyxl
import os
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook

if len(sys.argv) != 2:
    print("use: python catalog_names.py excel_file.xlsx")
    exit()

xlsx_file = sys.argv[1]
workbook = load_workbook(filename=xlsx_file)
sheet = workbook.active
name_library = {}


def dash(number):
    number = str(number)
    if len(number) == 9:
        return number[:3] + "-" + number[3:5] + "-" + number[5:]
    if len(number) == 11:
        return dash(number[:9]) + " " + number[9:]
    return number


def fill_library(file):
    soup = BeautifulSoup(file, "html.parser")
    naz = soup.find_all(name="td", attrs={"class": "naz"}, )
    nr = soup.find_all(name="td", attrs={"class": "nr"}, )
    for i in range(len(nr)):
        name_library[nr[i].string] = naz[i].string


def get_short_name(part_number):
    if name_library.get(part_number):
        return name_library.get(part_number)
    return ""


def fill_names_library_from_dir():
    file_pattern = re.compile(r'[0-9]{2}-[0-9]{2}.htm')
    for root, dirs, files in os.walk('.'):
        for file in files:
            if root.endswith('gb') and file_pattern.match(file):
                file_path = os.path.join(root, file)
                with open(file_path, encoding="windows-1250") as fp:
                    fill_library(fp)


fill_names_library_from_dir()

for row in sheet.iter_rows(max_col=3):
    part_number_dashed = dash(row[0].value)
    part_short_name = get_short_name(part_number_dashed)
    row[2].value = part_short_name
    workbook.save(xlsx_file)
